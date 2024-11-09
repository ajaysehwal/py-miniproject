from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

# Create a new workbook and add a worksheet
wb = Workbook()
ws = wb.active
ws.title = "Sheet1"

# Set column widths for better readability
column_widths = {
    'A': 90, 'B': 70, 'C': 85, 'D': 85, 'E': 85, 'F': 85, 'G': 85,
    'H': 85, 'I': 85, 'J': 85, 'K': 85, 'L': 85, 'M': 85, 'N': 85,
    'O': 85, 'P': 85, 'Q': 85
}

for col, width in column_widths.items():
    ws.column_dimensions[col].width = width / 6

# Header Row
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(bold=True, size=14, color="FFFFFF")
header_alignment = Alignment(horizontal="center", vertical="center")

# Title
ws.merge_cells("A1:Q1")
ws["A1"] = "DAILY TRADING JOURNAL"
ws["A1"].fill = header_fill
ws["A1"].font = header_font
ws["A1"].alignment = header_alignment

# Trade Details Section Header
section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
section_font = Font(bold=True, size=11)
section_alignment = Alignment(horizontal="center", vertical="center")

ws.merge_cells("A3:Q3")
ws["A3"] = "TRADE DETAILS"
ws["A3"].fill = section_fill
ws["A3"].font = section_font
ws["A3"].alignment = section_alignment

# Column Headers
header_fill2 = PatternFill(start_color="E7EEF5", end_color="E7EEF5", fill_type="solid")
header_font2 = Font(bold=True)
borders = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))

column_headers = [
    "Date", "Time", "Symbol", "Strategy", "Direction", "Entry Price", "Exit Price",
    "Position Size", "Stop Loss", "Take Profit", "Risk Per Trade ($)", "Reward:Risk",
    "P/L ($)", "P/L (%)", "Market Context", "Execution Grade", "Notes"
]

for col_num, header in enumerate(column_headers, 1):
    cell = ws.cell(row=4, column=col_num, value=header)
    cell.fill = header_fill2
    cell.font = header_font2
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = borders

# Data Validation
direction_validation = DataValidation(type="list", formula1='"LONG,SHORT"', showDropDown=True)
strategy_validation = DataValidation(type="list", formula1='"Breakout,Pull-back,Trend-following,Reversal,Swing"', showDropDown=True)
grade_validation = DataValidation(type="list", formula1='"A,B,C,D,F"', showDropDown=True)

ws.add_data_validation(direction_validation)
ws.add_data_validation(strategy_validation)
ws.add_data_validation(grade_validation)

direction_validation.add("E5:E1000")
strategy_validation.add("D5:D1000")
grade_validation.add("P5:P1000")

# Formulas for dynamic calculations
ws["A5"] = "=TODAY()"
ws["K5"] = "=H5*(F5-I5)"        # Risk Calculation
ws["L5"] = "=IF(G5>0,(G5-F5)/(F5-I5),\"\")"  # Reward:Risk Ratio
ws["M5"] = "=(G5-F5)*H5"        # P/L Calculation
ws["N5"] = "=M5/F5*100"         # P/L Percentage

# Daily Summary Section
ws.merge_cells("A15:B15")
ws["A15"] = "DAILY SUMMARY"
ws["A15"].fill = section_fill
ws["A15"].font = section_font
ws["A15"].alignment = section_alignment

daily_summary = [
    ("Total Trades:", "=COUNTA(C5:C1000)"),
    ("Winning Trades:", "=COUNTIF(M5:M1000,\">0\")"),
    ("Win Rate:", "=IF(B16>0,B17/B16,\"\")"),
    ("Total P/L:", "=SUM(M5:M1000)"),
    ("Avg. R:R Ratio:", "=AVERAGE(L5:L1000)"),
    ("Avg. P/L per Trade:", "=IF(B16>0, B18/B16, \"\")")
]

for row_num, (label, formula) in enumerate(daily_summary, start=16):
    ws[f"A{row_num}"] = label
    ws[f"B{row_num}"] = formula

# Weekly Metrics Section
ws.merge_cells("A22:B22")
ws["A22"] = "WEEKLY METRICS"
ws["A22"].fill = section_fill
ws["A22"].font = section_font
ws["A22"].alignment = section_alignment

weekly_metrics = [
    ("Best Trade:", "=MAX(M5:M1000)"),
    ("Worst Trade:", "=MIN(M5:M1000)"),
    ("Avg. Trade:", "=AVERAGE(M5:M1000)"),
    ("Total Commission:", "=SUM(K5:K1000)"),
    ("Net P/L:", "=B19-B26")
]

for row_num, (label, formula) in enumerate(weekly_metrics, start=23):
    ws[f"A{row_num}"] = label
    ws[f"B{row_num}"] = formula

# Monthly Summary Section
ws.merge_cells("A30:B30")
ws["A30"] = "MONTHLY SUMMARY"
ws["A30"].fill = section_fill
ws["A30"].font = section_font
ws["A30"].alignment = section_alignment

monthly_metrics = [
    ("Total Trades:", "=COUNTA(C5:C1000)"),
    ("Total Winning Trades:", "=COUNTIF(M5:M1000,\">0\")"),
    ("Total Losing Trades:", "=COUNTIF(M5:M1000,\"<0\")"),
    ("Monthly P/L:", "=SUM(M5:M1000)"),
    ("Monthly Win Rate:", "=IF(B32>0,B33/B32,\"\")")
]

for row_num, (label, formula) in enumerate(monthly_metrics, start=31):
    ws[f"A{row_num}"] = label
    ws[f"B{row_num}"] = formula

# Conditional Formatting for P/L
ws.conditional_formatting.add('M5:M1000', CellIsRule(operator='greaterThan', formula=['0'], fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")))
ws.conditional_formatting.add('M5:M1000', CellIsRule(operator='lessThan', formula=['0'], fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")))

# Save the workbook
wb.save("Daily_Trading_Journal.xlsx")
